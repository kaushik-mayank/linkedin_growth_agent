"""Validated parser for LinkedIn's 'Aggregate Analytics' XLSX export.

Ported from a hand-validated reference run against a real export file. The
quirks handled here are all confirmed from that real file — every number is
stored as a STRING, percentages read "< 1%", TOP POSTS holds two side-by-side
tables with the header on row 3, and any sheet (or the whole TOP POSTS table)
can be completely empty. Do not "clean up" this logic without re-testing
against a real export — see backend/tests/test_linkedin_parser.py.
"""
import io
import re
from datetime import datetime
from typing import Any

import openpyxl


def _num(v: Any) -> int | float:
    """LinkedIn stores numbers as STRINGS ('0', '1,162'). Coerce safely."""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(",", "")
    if s in ("", "-", "â"):
        return 0
    try:
        return float(s) if "." in s else int(s)
    except ValueError:
        return 0


def _pct(v: Any) -> float:
    """Percentages are strings like '2%' or '< 1%'."""
    if v is None:
        return 0.0
    s = str(v).strip().replace("%", "").replace("<", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def _date(v: Any) -> str | None:
    if isinstance(v, datetime):
        return v.date().isoformat()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(v).strip(), fmt).date().isoformat()
        except (ValueError, TypeError):
            pass
    return None


def parse(file_bytes: bytes) -> dict[str, Any]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheets = {s.upper(): s for s in wb.sheetnames}
    out: dict[str, Any] = {
        "period": {}, "discovery": {}, "daily": [], "top_posts": [],
        "followers": {}, "demographics": {}, "warnings": [],
    }

    # ---- DISCOVERY: date range lives in B1 as "7/10/2026 - 7/16/2026"
    if "DISCOVERY" in sheets:
        ws = wb[sheets["DISCOVERY"]]
        rng = str(ws.cell(1, 2).value or "")
        m = re.match(r"\s*(\d+/\d+/\d+)\s*-\s*(\d+/\d+/\d+)", rng)
        if m:
            out["period"] = {"start": _date(m.group(1)), "end": _date(m.group(2))}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                out["discovery"][str(row[0]).strip().lower().replace(" ", "_")] = _num(row[1])

    # ---- ENGAGEMENT: header row 1, then daily rows
    if "ENGAGEMENT" in sheets:
        for row in wb[sheets["ENGAGEMENT"]].iter_rows(min_row=2, values_only=True):
            d = _date(row[0])
            if d:
                out["daily"].append({"date": d, "impressions": _num(row[1]),
                                     "engagements": _num(row[2])})

    # ---- TOP POSTS: TWO side-by-side tables (A-C engagements, E-G impressions),
    #      header on row 3, may be completely empty.
    if "TOP POSTS" in sheets:
        ws = wb[sheets["TOP POSTS"]]
        merged: dict[str, dict[str, Any]] = {}
        for row in ws.iter_rows(min_row=4, values_only=True):
            if row and row[0]:
                merged.setdefault(str(row[0]), {}).update(
                    {"url": str(row[0]), "published": _date(row[1]), "engagements": _num(row[2])})
            if row and len(row) > 4 and row[4]:
                merged.setdefault(str(row[4]), {}).update(
                    {"url": str(row[4]), "published": _date(row[5]), "impressions": _num(row[6])})
        out["top_posts"] = list(merged.values())
        if not out["top_posts"]:
            out["warnings"].append("No posts in export period — treat as a silent week.")

    # ---- FOLLOWERS: 'Total followers on <date>' in A1/B1; header row 3
    if "FOLLOWERS" in sheets:
        ws = wb[sheets["FOLLOWERS"]]
        out["followers"]["total"] = _num(ws.cell(1, 2).value)
        daily = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            d = _date(row[0])
            if d:
                daily.append({"date": d, "new_followers": _num(row[1])})
        out["followers"]["daily"] = daily
        out["followers"]["new_this_period"] = sum(x["new_followers"] for x in daily)

    # ---- DEMOGRAPHICS: Category | Value | Percentage
    if "DEMOGRAPHICS" in sheets:
        for row in wb[sheets["DEMOGRAPHICS"]].iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] is not None:
                out["demographics"].setdefault(str(row[0]).strip(), []).append(
                    {"value": str(row[1]), "percentage": _pct(row[2])})

    tot_imp = sum(d["impressions"] for d in out["daily"])
    tot_eng = sum(d["engagements"] for d in out["daily"])
    out["totals"] = {"impressions": tot_imp, "engagements": tot_eng,
                     "engagement_rate": round(tot_eng / tot_imp * 100, 2) if tot_imp else 0.0,
                     "posts_published": len(out["top_posts"])}
    if tot_imp == 0:
        out["warnings"].append("Zero impressions: account was dormant this period.")
    return out
