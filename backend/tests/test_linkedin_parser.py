import io
from datetime import datetime

import openpyxl

from app.parser.linkedin_parser import _date, _num, _pct, parse


def _wb_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Landmine: every number is stored as a STRING, sometimes with thousands commas
# ---------------------------------------------------------------------------
def test_num_coerces_string_numbers():
    assert _num("0") == 0
    assert _num("1,162") == 1162
    assert _num(None) == 0
    assert _num("-") == 0
    assert _num("") == 0
    assert _num(42) == 42
    assert _num("3.5") == 3.5


# ---------------------------------------------------------------------------
# Landmine: percentages are text, including the unparseable-looking "< 1%"
# ---------------------------------------------------------------------------
def test_pct_handles_less_than_one_percent_without_crashing():
    assert _pct("2%") == 2.0
    assert _pct("< 1%") == 1.0
    assert _pct(None) == 0.0
    assert _pct("garbage") == 0.0


# ---------------------------------------------------------------------------
# Landmine: dates are M/D/YYYY strings, not date objects (but datetime objects
# also show up depending on how openpyxl reads the cell)
# ---------------------------------------------------------------------------
def test_date_handles_datetime_and_string_formats():
    assert _date(datetime(2026, 7, 10)) == "2026-07-10"
    assert _date("7/10/2026") == "2026-07-10"
    assert _date("2026-07-10") == "2026-07-10"
    assert _date("not a date") is None
    assert _date(None) is None


def _build_workbook(*, include_demographics=True, include_top_posts_data=True):
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "DISCOVERY"
    ws["B1"] = "7/10/2026 - 7/16/2026"
    ws.append(["Impressions", "0"])
    ws.append(["Members reached", "0"])

    ws = wb.create_sheet("ENGAGEMENT")
    ws.append(["Date", "Impressions", "Engagements"])  # header, ignored (min_row=2)
    # intentionally no daily rows -> zero impressions, a valid dormant week

    ws = wb.create_sheet("TOP POSTS")
    ws.append(["This is a note row"])
    ws.append([])
    ws.append(["Post URL", "Date", "Engagements", "", "Post URL", "Date", "Impressions"])
    if include_top_posts_data:
        ws.append([
            "https://post1", datetime(2026, 7, 11), "15", None,
            "https://post1", datetime(2026, 7, 11), "120",
        ])
        ws.append(["https://post2", datetime(2026, 7, 12), "8"])  # engagements only
        ws.append([None, None, None, None, "https://post3", datetime(2026, 7, 13), "50"])  # impressions only

    ws = wb.create_sheet("FOLLOWERS")
    ws["B1"] = "1,162"
    ws.append([])
    ws.append(["Date", "New followers"])
    ws.append(["7/12/2026", "5"])
    ws.append(["7/14/2026", "7"])

    if include_demographics:
        ws = wb.create_sheet("DEMOGRAPHICS")
        ws.append(["Category", "Value", "Percentage"])
        ws.append(["Company", "🦸 Data Heroes", "20%"])
        ws.append(["Location", "Delhi", "38%"])
        ws.append(["Company size", "1,001-5,000", "< 1%"])
        ws.append(["Seniority", "Senior", "45%"])
        ws.append(["Job title", "Data Analyst", "12%"])
        ws.append(["Industry", "IT Services", "30%"])

    return wb


def test_top_posts_merges_two_side_by_side_tables_by_url():
    result = parse(_wb_bytes(_build_workbook()))
    by_url = {p["url"]: p for p in result["top_posts"]}

    assert by_url["https://post1"]["engagements"] == 15
    assert by_url["https://post1"]["impressions"] == 120
    assert by_url["https://post2"]["engagements"] == 8
    assert "impressions" not in by_url["https://post2"]
    assert by_url["https://post3"]["impressions"] == 50
    assert "engagements" not in by_url["https://post3"]


def test_top_posts_completely_empty_is_a_valid_silent_week():
    result = parse(_wb_bytes(_build_workbook(include_top_posts_data=False)))
    assert result["top_posts"] == []
    assert any("silent week" in w for w in result["warnings"])


def test_demographics_handles_emoji_values():
    result = parse(_wb_bytes(_build_workbook()))
    company = result["demographics"]["Company"]
    assert company[0]["value"] == "🦸 Data Heroes"
    assert company[0]["percentage"] == 20.0
    # "< 1%" must not crash and must produce a number
    size = result["demographics"]["Company size"]
    assert size[0]["percentage"] == 1.0


def test_missing_sheet_does_not_crash_company_page_export():
    result = parse(_wb_bytes(_build_workbook(include_demographics=False)))
    assert result["demographics"] == {}
    assert result["totals"]["impressions"] == 0


def test_full_parse_matches_expected_dormant_week_shape():
    result = parse(_wb_bytes(_build_workbook(include_top_posts_data=False)))

    assert result["period"] == {"start": "2026-07-10", "end": "2026-07-16"}
    assert result["followers"]["total"] == 1162
    assert result["followers"]["new_this_period"] == 12
    assert result["totals"]["impressions"] == 0
    assert result["totals"]["posts_published"] == 0
    assert len(result["demographics"]) == 6
    assert any("dormant" in w for w in result["warnings"])
    assert any("silent week" in w for w in result["warnings"])
